from tkinter import *
import math, random, os
from tkinter import messagebox
from datetime import datetime
from openpyxl import load_workbook


class Bill_App:
    def __init__(self, root):
        self.root = root
        self.root.geometry("1350x700+0+0")
        self.root.title("Billing Software")
        self.root.resizable(False, False)
        bg_color = "#074463"
        title = Label(
            self.root,
            text="Billing Software",
            bd=12,
            relief=GROOVE,
            bg=bg_color,
            fg="white",
            font=("times new roman", 30, "bold"),
            pady=2,
        ).pack(fill=X)

        # ====Variables====

        # ====Cosmetics====
        self.soap = IntVar()
        self.face_cream = IntVar()
        self.face_wash = IntVar()
        self.spray = IntVar()
        self.gel = IntVar()
        self.loshan = IntVar()

        # ====Grocery====
        self.rice = IntVar()
        self.food_oil = IntVar()
        self.daal = IntVar()
        self.wheat = IntVar()
        self.sugar = IntVar()
        self.tea = IntVar()

        # ====Cold Drink====
        self.maza = IntVar()
        self.coke = IntVar()
        self.frooti = IntVar()
        self.thubs_up = IntVar()
        self.limca = IntVar()
        self.sprite = IntVar()

        # ====Total Product Price & Tax Variable====

        self.cosmetic_price = StringVar()
        self.grocery_price = StringVar()
        self.cold_drink_price = StringVar()

        self.cosmetic_tax = StringVar()
        self.grocery_tax = StringVar()
        self.cold_drink_tax = StringVar()

        # ====Customer====
        self.c_name = StringVar()
        self.c_phone = StringVar()

        self.bill_no = StringVar()
        self.x = datetime.now()
        self.bill_no.set(str(self.x.strftime("%H%M%S")))

        self.search_bill = StringVar()

        # ==== Customer Detail Frame====
        F1 = LabelFrame(
            self.root,
            bd=10,
            relief=GROOVE,
            text="Customer Details",
            font=("times new roman", 15, "bold"),
            fg="gold",
            bg=bg_color,
        )
        F1.place(x=0, y=73, relwidth=1)

        cname_lbl = Label(
            F1,
            text="Customer Name",
            bg=bg_color,
            fg="white",
            font=("times new roman", 18, "bold"),
        ).grid(row=0, column=0, padx=20, pady=5)
        cname_txt = Entry(
            F1, width=15, textvariable=self.c_name, font="arial 15", bd=7, relief=SUNKEN
        ).grid(row=0, column=1, padx=10, pady=5)

        cphn_lbl = Label(
            F1,
            text="Phone No.",
            bg=bg_color,
            fg="white",
            font=("times new roman", 18, "bold"),
        ).grid(row=0, column=2, padx=20, pady=5)
        cphn_txt = Entry(
            F1,
            width=15,
            textvariable=self.c_phone,
            font="arial 15",
            bd=7,
            relief=SUNKEN,
        ).grid(row=0, column=3, padx=10, pady=5)

        c_bill_lbl = Label(
            F1,
            text="Bill Number",
            bg=bg_color,
            fg="white",
            font=("times new roman", 18, "bold"),
        ).grid(row=0, column=4, padx=20, pady=5)
        c_bill_txt = Entry(
            F1,
            width=15,
            textvariable=self.search_bill,
            font="arial 15",
            bd=7,
            relief=SUNKEN,
        ).grid(row=0, column=5, padx=10, pady=5)

        bill_btn = Button(
            F1,
            text="Search",
            command=self.find_bill,
            width=10,
            bd=7,
            font="arial 12 bold",
        ).grid(row=0, column=6, pady=10, padx=10)

        # ====Cosmetics Frame====
        F2 = LabelFrame(
            self.root,
            bd=10,
            relief=GROOVE,
            text="Cosmetics",
            font=("times new roman", 15, "bold"),
            fg="gold",
            bg=bg_color,
        )
        F2.place(x=0, y=170, width=325, height=380)

        bath_lbl = Label(
            F2,
            text="Bath Soap",
            font=("times new roman", 16, "bold"),
            bg=bg_color,
            fg="lightgreen",
        ).grid(row=0, column=0, padx=10, pady=10, sticky="w")
        bath_txt = Entry(
            F2,
            width=10,
            textvariable=self.soap,
            font=("times new roman", 16, "bold"),
            bd=5,
            relief=SUNKEN,
        ).grid(row=0, column=1, padx=10, pady=10)

        Face_cream_lbl = Label(
            F2,
            text="Face Cream",
            font=("times new roman", 16, "bold"),
            bg=bg_color,
            fg="lightgreen",
        ).grid(row=1, column=0, padx=10, pady=10, sticky="w")
        Face_cream_txt = Entry(
            F2,
            width=10,
            textvariable=self.face_cream,
            font=("times new roman", 16, "bold"),
            bd=5,
            relief=SUNKEN,
        ).grid(row=1, column=1, padx=10, pady=10)

        Face_w_lbl = Label(
            F2,
            text="Face Wash",
            font=("times new roman", 16, "bold"),
            bg=bg_color,
            fg="lightgreen",
        ).grid(row=2, column=0, padx=10, pady=10, sticky="w")
        Face_w_txt = Entry(
            F2,
            width=10,
            textvariable=self.face_wash,
            font=("times new roman", 16, "bold"),
            bd=5,
            relief=SUNKEN,
        ).grid(row=2, column=1, padx=10, pady=10)

        Hair_s_lbl = Label(
            F2,
            text="Hair Spray",
            font=("times new roman", 16, "bold"),
            bg=bg_color,
            fg="lightgreen",
        ).grid(row=3, column=0, padx=10, pady=10, sticky="w")
        Hair_s_txt = Entry(
            F2,
            width=10,
            textvariable=self.spray,
            font=("times new roman", 16, "bold"),
            bd=5,
            relief=SUNKEN,
        ).grid(row=3, column=1, padx=10, pady=10)

        Hair_g_lbl = Label(
            F2,
            text="Hair Gel",
            font=("times new roman", 16, "bold"),
            bg=bg_color,
            fg="lightgreen",
        ).grid(row=4, column=0, padx=10, pady=10, sticky="w")
        Hair_g_txt = Entry(
            F2,
            width=10,
            textvariable=self.gel,
            font=("times new roman", 16, "bold"),
            bd=5,
            relief=SUNKEN,
        ).grid(row=4, column=1, padx=10, pady=10)

        Body_lbl = Label(
            F2,
            text="Body Lotion",
            font=("times new roman", 16, "bold"),
            bg=bg_color,
            fg="lightgreen",
        ).grid(row=5, column=0, padx=10, pady=10, sticky="w")
        Body_txt = Entry(
            F2,
            width=10,
            textvariable=self.loshan,
            font=("times new roman", 16, "bold"),
            bd=5,
            relief=SUNKEN,
        ).grid(row=5, column=1, padx=10, pady=10)

        # ====Grocery Frame====

        F3 = LabelFrame(
            self.root,
            bd=10,
            relief=GROOVE,
            text="Grocery",
            font=("times new roman", 15, "bold"),
            fg="gold",
            bg=bg_color,
        )
        F3.place(x=325, y=170, width=325, height=380)

        g1_lbl = Label(
            F3,
            text="Rice",
            font=("times new roman", 16, "bold"),
            bg=bg_color,
            fg="lightgreen",
        ).grid(row=0, column=0, padx=10, pady=10, sticky="w")
        g1_txt = Entry(
            F3,
            width=10,
            textvariable=self.rice,
            font=("times new roman", 16, "bold"),
            bd=5,
            relief=SUNKEN,
        ).grid(row=0, column=1, padx=10, pady=10)

        g2_lbl = Label(
            F3,
            text="Food Oil",
            font=("times new roman", 16, "bold"),
            bg=bg_color,
            fg="lightgreen",
        ).grid(row=1, column=0, padx=10, pady=10, sticky="w")
        g2_txt = Entry(
            F3,
            width=10,
            textvariable=self.food_oil,
            font=("times new roman", 16, "bold"),
            bd=5,
            relief=SUNKEN,
        ).grid(row=1, column=1, padx=10, pady=10)

        g3_lbl = Label(
            F3,
            text="Daal",
            font=("times new roman", 16, "bold"),
            bg=bg_color,
            fg="lightgreen",
        ).grid(row=2, column=0, padx=10, pady=10, sticky="w")
        g3_txt = Entry(
            F3,
            width=10,
            textvariable=self.daal,
            font=("times new roman", 16, "bold"),
            bd=5,
            relief=SUNKEN,
        ).grid(row=2, column=1, padx=10, pady=10)

        g4_lbl = Label(
            F3,
            text="Wheat",
            font=("times new roman", 16, "bold"),
            bg=bg_color,
            fg="lightgreen",
        ).grid(row=3, column=0, padx=10, pady=10, sticky="w")
        g4_txt = Entry(
            F3,
            width=10,
            textvariable=self.wheat,
            font=("times new roman", 16, "bold"),
            bd=5,
            relief=SUNKEN,
        ).grid(row=3, column=1, padx=10, pady=10)

        g5_lbl = Label(
            F3,
            text="Sugar",
            font=("times new roman", 16, "bold"),
            bg=bg_color,
            fg="lightgreen",
        ).grid(row=4, column=0, padx=10, pady=10, sticky="w")
        g5_txt = Entry(
            F3,
            width=10,
            textvariable=self.sugar,
            font=("times new roman", 16, "bold"),
            bd=5,
            relief=SUNKEN,
        ).grid(row=4, column=1, padx=10, pady=10)

        g6_lbl = Label(
            F3,
            text="Tea",
            font=("times new roman", 16, "bold"),
            bg=bg_color,
            fg="lightgreen",
        ).grid(row=5, column=0, padx=10, pady=10, sticky="w")
        g6_txt = Entry(
            F3,
            width=10,
            textvariable=self.tea,
            font=("times new roman", 16, "bold"),
            bd=5,
            relief=SUNKEN,
        ).grid(row=5, column=1, padx=10, pady=10)

        # ====Cold Drink====

        F4 = LabelFrame(
            self.root,
            bd=10,
            relief=GROOVE,
            text="Cold Drinks",
            font=("times new roman", 15, "bold"),
            fg="gold",
            bg=bg_color,
        )
        F4.place(x=650, y=170, width=325, height=380)

        c1_lbl = Label(
            F4,
            text="Maza",
            font=("times new roman", 16, "bold"),
            bg=bg_color,
            fg="lightgreen",
        ).grid(row=0, column=0, padx=10, pady=10, sticky="w")
        c1_txt = Entry(
            F4,
            width=10,
            textvariable=self.maza,
            font=("times new roman", 16, "bold"),
            bd=5,
            relief=SUNKEN,
        ).grid(row=0, column=1, padx=10, pady=10)

        c2_lbl = Label(
            F4,
            text="Coke",
            font=("times new roman", 16, "bold"),
            bg=bg_color,
            fg="lightgreen",
        ).grid(row=1, column=0, padx=10, pady=10, sticky="w")
        c2_txt = Entry(
            F4,
            width=10,
            textvariable=self.coke,
            font=("times new roman", 16, "bold"),
            bd=5,
            relief=SUNKEN,
        ).grid(row=1, column=1, padx=10, pady=10)

        c3_lbl = Label(
            F4,
            text="Frooti",
            font=("times new roman", 16, "bold"),
            bg=bg_color,
            fg="lightgreen",
        ).grid(row=2, column=0, padx=10, pady=10, sticky="w")
        c3_txt = Entry(
            F4,
            width=10,
            textvariable=self.frooti,
            font=("times new roman", 16, "bold"),
            bd=5,
            relief=SUNKEN,
        ).grid(row=2, column=1, padx=10, pady=10)

        c4_lbl = Label(
            F4,
            text="Thumbs Up",
            font=("times new roman", 16, "bold"),
            bg=bg_color,
            fg="lightgreen",
        ).grid(row=3, column=0, padx=10, pady=10, sticky="w")
        c4_txt = Entry(
            F4,
            width=10,
            textvariable=self.thubs_up,
            font=("times new roman", 16, "bold"),
            bd=5,
            relief=SUNKEN,
        ).grid(row=3, column=1, padx=10, pady=10)

        c5_lbl = Label(
            F4,
            text="Limca",
            font=("times new roman", 16, "bold"),
            bg=bg_color,
            fg="lightgreen",
        ).grid(row=4, column=0, padx=10, pady=10, sticky="w")
        c5_txt = Entry(
            F4,
            width=10,
            textvariable=self.limca,
            font=("times new roman", 16, "bold"),
            bd=5,
            relief=SUNKEN,
        ).grid(row=4, column=1, padx=10, pady=10)

        c6_lbl = Label(
            F4,
            text="Sprite",
            font=("times new roman", 16, "bold"),
            bg=bg_color,
            fg="lightgreen",
        ).grid(row=5, column=0, padx=10, pady=10, sticky="w")
        c6_txt = Entry(
            F4,
            width=10,
            textvariable=self.sprite,
            font=("times new roman", 16, "bold"),
            bd=5,
            relief=SUNKEN,
        ).grid(row=5, column=1, padx=10, pady=10)

        # ====Bill Area====

        F5 = Frame(self.root, bd=10, relief=GROOVE)
        F5.place(x=970, y=170, width=380, height=380)
        bill_title = Label(
            F5, text="Bill Area", font="arial 15 bold", bd=7, relief=GROOVE
        ).pack(fill=X)
        scroll_y = Scrollbar(F5, orient=VERTICAL)
        self.txtarea = Text(F5, yscrollcommand=scroll_y.set)
        scroll_y.pack(side=RIGHT, fill=Y)
        scroll_y.config(command=self.txtarea.yview)
        self.txtarea.pack(fill=BOTH, expand=1)

        # ====Button Frame====

        F6 = LabelFrame(
            self.root,
            bd=10,
            relief=GROOVE,
            text="Bill Menu",
            font=("times new roman", 15, "bold"),
            fg="gold",
            bg=bg_color,
        )
        F6.place(x=0, y=550, relwidth=1, height=150)

        m1_lbl = Label(
            F6,
            text="Total Cosmetic Price",
            bg=bg_color,
            fg="white",
            font=("times new roman", 14, "bold"),
        ).grid(row=0, column=0, padx=20, pady=1, sticky="w")
        m1_txt = Entry(
            F6,
            width=18,
            textvariable=self.cosmetic_price,
            font="arial 10 bold",
            bd=7,
            relief=SUNKEN,
        ).grid(row=0, column=1, padx=10, pady=1)

        m2_lbl = Label(
            F6,
            text="Total Grocery Price",
            bg=bg_color,
            fg="white",
            font=("times new roman", 14, "bold"),
        ).grid(row=1, column=0, padx=20, pady=1, sticky="w")
        m2_txt = Entry(
            F6,
            width=18,
            textvariable=self.grocery_price,
            font="arial 10 bold",
            bd=7,
            relief=SUNKEN,
        ).grid(row=1, column=1, padx=10, pady=1)

        m3_lbl = Label(
            F6,
            text="Total Cold Drinks Price",
            bg=bg_color,
            fg="white",
            font=("times new roman", 14, "bold"),
        ).grid(row=2, column=0, padx=20, pady=1, sticky="w")
        m3_txt = Entry(
            F6,
            width=18,
            textvariable=self.cold_drink_price,
            font="arial 10 bold",
            bd=7,
            relief=SUNKEN,
        ).grid(row=2, column=1, padx=10, pady=1)

        c1_lbl = Label(
            F6,
            text="Cosmetic Tax",
            bg=bg_color,
            fg="white",
            font=("times new roman", 14, "bold"),
        ).grid(row=0, column=2, padx=20, pady=1, sticky="w")
        c1_txt = Entry(
            F6,
            width=18,
            textvariable=self.cosmetic_tax,
            font="arial 10 bold",
            bd=7,
            relief=SUNKEN,
        ).grid(row=0, column=3, padx=10, pady=1)

        c2_lbl = Label(
            F6,
            text="Grocery Tax",
            bg=bg_color,
            fg="white",
            font=("times new roman", 14, "bold"),
        ).grid(row=1, column=2, padx=20, pady=1, sticky="w")
        c2_txt = Entry(
            F6,
            width=18,
            textvariable=self.grocery_tax,
            font="arial 10 bold",
            bd=7,
            relief=SUNKEN,
        ).grid(row=1, column=3, padx=10, pady=1)

        c3_lbl = Label(
            F6,
            text="Cold Drinks Tax",
            bg=bg_color,
            fg="white",
            font=("times new roman", 14, "bold"),
        ).grid(row=2, column=2, padx=20, pady=1, sticky="w")
        c3_txt = Entry(
            F6,
            width=18,
            textvariable=self.cold_drink_tax,
            font="arial 10 bold",
            bd=7,
            relief=SUNKEN,
        ).grid(row=2, column=3, padx=10, pady=1)

        # ====Button Frame====
        btn_F = Frame(F6, bd=7, relief=GROOVE)
        btn_F.place(x=750, width=580, height=105)

        total_btn = Button(
            btn_F,
            command=self.total,
            text="Total",
            bg="cadetblue",
            fg="white",
            pady=15,
            bd=2,
            width=10,
            font="arial 15 bold",
        ).grid(row=0, column=0, padx=6, pady=7)

        Gbill_btn = Button(
            btn_F,
            text="Generate",
            command=self.bill_area,
            bg="cadetblue",
            fg="white",
            pady=15,
            bd=2,
            width=10,
            font="arial 15 bold",
        ).grid(row=0, column=1, padx=5, pady=7)

        Clear_btn = Button(
            btn_F,
            text="Clear",
            bg="cadetblue",
            command=self.clear_data,
            fg="white",
            pady=15,
            bd=2,
            width=10,
            font="arial 15 bold",
        ).grid(row=0, column=2, padx=5, pady=7)

        Exit_btn = Button(
            btn_F,
            text="Exit",
            bg="cadetblue",
            command=self.Exit_app,
            fg="white",
            pady=15,
            bd=2,
            width=10,
            font="arial 15 bold",
        ).grid(row=0, column=3, padx=5, pady=7)
        self.welcome_bill()

    def total(self):
        self.c_s_p = self.soap.get() * 40
        self.c_fc_p = self.face_cream.get() * 120
        self.c_fw_p = self.face_wash.get() * 60
        self.c_hs_p = self.spray.get() * 180
        self.c_hg_p = self.gel.get() * 140
        self.c_bl_p = self.loshan.get() * 180
        self.total_cosmetic_price = float(
            self.c_s_p
            + self.c_fc_p
            + self.c_fw_p
            + self.c_hs_p
            + self.c_hg_p
            + self.c_bl_p
        )
        self.cosmetic_price.set("Rs " + str(self.total_cosmetic_price))
        self.c_tax = round(self.total_cosmetic_price * 0.05, 2)
        self.cosmetic_tax.set("Rs " + str(self.c_tax))

        self.g_r_p = self.rice.get() * 80
        self.g_f_p = self.food_oil.get() * 180
        self.g_d_p = self.daal.get() * 60
        self.g_w_p = self.wheat.get() * 240
        self.g_s_p = self.sugar.get() * 45
        self.g_t_p = self.tea.get() * 150
        self.total_grocery_price = float(
            self.g_r_p + self.g_f_p + self.g_d_p + self.g_w_p + self.g_s_p + self.g_t_p
        )
        self.grocery_price.set("Rs " + str(self.total_grocery_price))
        self.g_tax = round(self.total_grocery_price * 0.1, 2)
        self.grocery_tax.set("Rs " + str(self.g_tax))

        self.cd_m_p = self.maza.get() * 60
        self.cd_c_p = self.coke.get() * 60
        self.cd_f_p = self.frooti.get() * 50
        self.cd_t_p = self.thubs_up.get() * 45
        self.cd_l_p = self.limca.get() * 40
        self.cd_s_p = self.sprite.get() * 60
        self.total_cold_drink_price = float(
            self.cd_m_p
            + self.cd_c_p
            + self.cd_f_p
            + self.cd_t_p
            + self.cd_l_p
            + self.cd_s_p
        )
        self.cold_drink_price.set("Rs " + str(self.total_cold_drink_price))
        self.d_tax = round(self.total_cold_drink_price * 0.1, 2)
        self.cold_drink_tax.set("Rs " + str(self.d_tax))

        self.Total_bill = float(
            self.total_cosmetic_price
            + self.total_grocery_price
            + self.total_cold_drink_price
            + self.c_tax
            + self.g_tax
            + self.d_tax
        )

    def welcome_bill(self):
        self.txtarea.delete("1.0", END)
        self.txtarea.insert(END, "\t   Welcome Sarthak Retail\n")
        self.txtarea.insert(END, f" \t        {self.x.date()}\n")
        self.txtarea.insert(END, f"\n==========================================")
        self.txtarea.insert(END, f"\n Bill Number : {self.bill_no.get()}")
        self.txtarea.insert(END, f"\n Customer Name : {self.c_name.get()}")
        self.txtarea.insert(END, f"\n Phone Number : {self.c_phone.get()}")
        self.txtarea.insert(END, f"\n==========================================")
        self.txtarea.insert(END, f"\n Products \t\t QTY \t \t Price")
        self.txtarea.insert(END, f"\n==========================================")

    def bill_area(self):
        if self.c_name.get() == "" or self.c_phone.get() == "":
            messagebox.showerror("Error", "Customer Details are must")
        elif (
            self.cosmetic_price.get() == "Rs 0.0"
            and self.grocery_price.get() == "Rs 0.0"
            and self.cold_drink_price.get() == "Rs 0.0"
        ):
            messagebox.showerror("Error", "No Product Purchased")
        else:
            self.welcome_bill()

            # ====Cosmetic====
            if self.soap.get() != 0:
                self.txtarea.insert(
                    END, f"\n Bath Soap \t\t {self.soap.get()} \t\t {self.c_s_p}"
                )
            if self.face_cream.get() != 0:
                self.txtarea.insert(
                    END,
                    f"\n Face Cream \t\t {self.face_cream.get()} \t\t {self.c_fc_p}",
                )
            if self.face_wash.get() != 0:
                self.txtarea.insert(
                    END, f"\n Face Wash \t\t {self.face_wash.get()} \t\t {self.c_fw_p}"
                )
            if self.spray.get() != 0:
                self.txtarea.insert(
                    END, f"\n Hair Spray \t\t {self.spray.get()} \t\t {self.c_hs_p}"
                )
            if self.gel.get() != 0:
                self.txtarea.insert(
                    END, f"\n Hair Gel \t\t {self.gel.get()} \t\t {self.c_hg_p}"
                )
            if self.loshan.get() != 0:
                self.txtarea.insert(
                    END, f"\n Body Loashan \t\t {self.loshan.get()} \t\t {self.c_bl_p}"
                )

            # ====Grocery====
            if self.rice.get() != 0:
                self.txtarea.insert(
                    END, f"\n Rice \t\t {self.rice.get()} \t\t {self.g_r_p}"
                )
            if self.food_oil.get() != 0:
                self.txtarea.insert(
                    END, f"\n Food Oil \t\t {self.food_oil.get()} \t\t {self.g_f_p}"
                )
            if self.daal.get() != 0:
                self.txtarea.insert(
                    END, f"\n Daal \t\t {self.daal.get()} \t\t {self.g_d_p}"
                )
            if self.wheat.get() != 0:
                self.txtarea.insert(
                    END, f"\n Wheat \t\t {self.wheat.get()} \t\t {self.g_w_p}"
                )
            if self.sugar.get() != 0:
                self.txtarea.insert(
                    END, f"\n Sugar \t\t {self.sugar.get()} \t\t {self.g_s_p}"
                )
            if self.tea.get() != 0:
                self.txtarea.insert(
                    END, f"\n Loshan \t\t {self.tea.get()} \t\t {self.g_t_p}"
                )

            # ====Cold Drink====
            if self.maza.get() != 0:
                self.txtarea.insert(
                    END, f"\n Maza \t\t {self.maza.get()} \t\t {self.g_r_p}"
                )
            if self.coke.get() != 0:
                self.txtarea.insert(
                    END, f"\n Coke \t\t {self.coke.get()} \t\t {self.cd_c_p}"
                )
            if self.frooti.get() != 0:
                self.txtarea.insert(
                    END, f"\n Frooti \t\t {self.frooti.get()} \t\t {self.cd_f_p}"
                )
            if self.thubs_up.get() != 0:
                self.txtarea.insert(
                    END, f"\n Thumbs Up \t\t {self.thubs_up.get()} \t\t {self.cd_t_p}"
                )
            if self.limca.get() != 0:
                self.txtarea.insert(
                    END, f"\n Limca \t\t {self.limca.get()} \t\t {self.cd_l_p}"
                )
            if self.sprite.get() != 0:
                self.txtarea.insert(
                    END, f"\n Sprite \t\t {self.sprite.get()} \t\t {self.cd_s_p}"
                )

            # ====Tax Print====
            self.txtarea.insert(END, f"\n------------------------------------------")
            if self.cosmetic_tax.get() != "Rs 0.0":
                self.txtarea.insert(
                    END, f"\n Cosmetic Tax \t\t\t\t{self.cosmetic_tax.get()}"
                )
            if self.grocery_tax.get() != "Rs 0.0":
                self.txtarea.insert(
                    END, f"\n Grocery Tax \t\t\t\t{self.grocery_tax.get()}"
                )
            if self.cold_drink_tax.get() != "Rs 0.0":
                self.txtarea.insert(
                    END, f"\n Cold Drink Tax \t\t\t\t{self.cold_drink_tax.get()}"
                )

            self.txtarea.insert(END, f"\n------------------------------------------")
            self.txtarea.insert(END, f"\n Total Bill \t\t\t\tRs {self.Total_bill}")
            self.txtarea.insert(END, f"\n------------------------------------------")
            self.save_bill()

    def save_bill(self):
        op = messagebox.askyesno("Save Bill", "Do you want to save the Bill?")
        if op > 0:
            self.bill_data = self.txtarea.get("1.0", END)
            f1 = open("bills/" + str(self.bill_no.get()) + ".txt", "w")
            f1.write(self.bill_data)
            f1.close()
            messagebox.showinfo(
                "Saved", f"Bill no. : {self.bill_no.get()} saved successfully"
            )

            # ==== Excel Work ====

            workbook_name = "DataBase/" + datetime.now().strftime("%Y") + ".xlsx"
            workbook = load_workbook(filename=workbook_name)
            sheet_name = datetime.now().strftime("%B")

            if sheet_name not in workbook.sheetnames:
                source_sheet = workbook["Do_Not_Alter"]
                new_sheet = workbook.copy_worksheet(source_sheet)
                new_sheet.title = sheet_name

            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                if all(cell.value is None for cell in row):
                    first_empty_row = row[0].row
                    break
            sheet.cell(row=first_empty_row, column=1).value = self.bill_no.get()
            sheet.cell(row=first_empty_row, column=2).value = datetime.now().strftime(
                "%d%m%y"
            )
            sheet.cell(row=first_empty_row, column=3).value = self.c_name.get()
            sheet.cell(row=first_empty_row, column=4).value = self.c_phone.get()
            sheet.cell(row=first_empty_row, column=5).value = (
                self.c_tax + self.g_tax + self.d_tax
            )
            sheet.cell(row=first_empty_row, column=6).value = self.Total_bill
            workbook.save(filename=workbook_name)

        else:
            return

    def find_bill(self):
        present = "no"
        for i in os.listdir("bills/"):
            if i.split(".")[0] == self.search_bill.get():
                f1 = open(f"bills/{i}", "r")
                self.txtarea.delete("1.0", END)
                for d in f1:
                    self.txtarea.insert(END, d)
                f1.close()
                present = "yes"
        if present == "no":
            messagebox.showerror("Error", "Invalid Bill Number")

    def clear_data(self):
        op = messagebox.askyesno("Clear", "Do you really want to clear?")
        if op > 0:
            # ====Cosmetics====
            self.soap.set(0)
            self.face_cream.set(0)
            self.face_wash.set(0)
            self.spray.set(0)
            self.gel.set(0)
            self.loshan.set(0)

            # ====Grocery====
            self.rice.set(0)
            self.food_oil.set(0)
            self.daal.set(0)
            self.wheat.set(0)
            self.sugar.set(0)
            self.tea.set(0)

            # ====Cold Drink====
            self.maza.set(0)
            self.coke.set(0)
            self.frooti.set(0)
            self.thubs_up.set(0)
            self.limca.set(0)
            self.sprite.set(0)

            # ====Total Product Price & Tax Variable====

            self.cosmetic_price.set("")
            self.grocery_price.set("")
            self.cold_drink_price.set("")

            self.cosmetic_tax.set("")
            self.grocery_tax.set("")
            self.cold_drink_tax.set("")

            # ====Customer====
            self.c_name.set("")
            self.c_phone.set("")

            self.bill_no.set("")
            self.x = datetime.now()
            self.bill_no.set(str(self.x.strftime("%H%M%S")))

            self.search_bill.set("")
            self.welcome_bill()

    def Exit_app(self):
        op = messagebox.askyesno("Exit", "Do you really want to exit?")
        if op > 0:
            self.root.destroy()


root = Tk()
obj = Bill_App(root)
root.mainloop()
